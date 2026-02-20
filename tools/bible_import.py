#!/usr/bin/env python3
"""bible_import.py

Deterministic XLSX -> normalized JSON exporter for the web runtime.

Outputs (schema-stable):
  data/generated/creatures.json
  data/generated/weapons.json
  data/generated/weapon_families.json
  data/generated/stages.json
  data/generated/props.json
  data/generated/bosses.json
  data/generated/evolution/*.json
  data/generated/vfx.json
  data/generated/sfx.json
  data/generated/ui_style_guide.json
  data/generated/hud_layout.json
  data/generated/validation_report.json

Non-destructive guarantee:
  - Never renames IDs
  - Missing/inconsistent data becomes warnings + safe fallbacks
"""

from __future__ import annotations

import argparse
import json
import os
import re
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


def _stable(obj: Any) -> Any:
    if isinstance(obj, list):
        # Stable sort for lists of objects with id.
        if obj and all(isinstance(x, dict) and "id" in x for x in obj):
            return sorted([_stable(x) for x in obj], key=lambda d: str(d.get("id", "")))
        return [_stable(x) for x in obj]
    if isinstance(obj, dict):
        return {k: _stable(obj[k]) for k in sorted(obj.keys())}
    return obj


def write_json(path: str, data: Any) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(_stable(data), f, indent=2)
        f.write("\n")


def find_header_row(ws, required: str = "id", scan_rows: int = 25) -> Optional[int]:
    required = required.strip().lower()
    for r in range(1, scan_rows + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        norm = [str(v).strip().lower() if v is not None else "" for v in vals]
        if required in norm:
            return r
    return None


def read_table(ws, id_col_names: Tuple[str, ...] = ("id",), scan_rows: int = 25) -> List[Dict[str, Any]]:
    # Find header row by any of the id column names.
    header_row = None
    chosen_id = None
    for cand in id_col_names:
        hr = find_header_row(ws, cand, scan_rows)
        if hr is not None:
            header_row = hr
            chosen_id = cand
            break
    if header_row is None:
        return []

    headers: List[str] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        h = str(v).strip() if v is not None else ""
        headers.append(h)

    out: List[Dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        row: Dict[str, Any] = {}
        empty = True
        for c, h in enumerate(headers, start=1):
            if not h:
                continue
            v = ws.cell(r, c).value
            if v is not None and v != "":
                empty = False
            row[h] = v
        if empty:
            continue

        # Require id column to exist and be non-empty.
        id_val = row.get(chosen_id) or row.get(chosen_id.capitalize())
        if id_val is None or str(id_val).strip() == "":
            continue
        out.append(row)
    return out


def parse_palette_tokens(text: str) -> Dict[str, str]:
    # e.g. "Background #141018; PanelBase #2A1633; ..."
    tokens: Dict[str, str] = {}
    for name, hexv in re.findall(r"([A-Za-z0-9_()]+)\s*(#[0-9A-Fa-f]{6})", text or ""):
        key = re.sub(r"[^A-Za-z0-9_]+", "_", name).strip("_")
        tokens[key] = hexv.lower()
    return tokens


def parse_hud_boxes(text: str) -> Dict[str, Any]:
    # From HUD_Spec "Coordinates 640×360" row.
    # Example: "HP x16 y16 w180 h20; ..."
    boxes = {}
    for part in (text or "").split(";"):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"^(?P<name>[A-Za-z0-9_ ]+)\s+(?P<rest>.*)$", part)
        if not m:
            continue
        name = m.group("name").strip().lower().replace(" ", "_")
        rest = m.group("rest")
        nums = dict(re.findall(r"\b([xywh])\s*([0-9]+)\b", rest))
        if nums:
            boxes[name] = {k: int(v) for k, v in nums.items()}
        else:
            boxes[name] = {"spec": rest}
    return {"base": {"w": 640, "h": 360}, "boxes": boxes}


@dataclass
class Validation:
    missing_references: List[Dict[str, Any]]
    duplicate_ids: List[Dict[str, Any]]
    evolution_errors: List[Dict[str, Any]]
    missing_assets: List[str]


def validate_all(
    creatures: List[Dict[str, Any]],
    weapons: List[Dict[str, Any]],
    stages: List[Dict[str, Any]],
    bosses: List[Dict[str, Any]],
    evo_nodes: List[Dict[str, Any]],
    evo_prereqs: List[Dict[str, Any]],
    assets_root: str,
) -> Validation:
    missing_refs: List[Dict[str, Any]] = []
    dupes: List[Dict[str, Any]] = []
    evo_errs: List[Dict[str, Any]] = []
    missing_assets: List[str] = []

    def check_dupes(items: List[Dict[str, Any]], key: str, kind: str) -> None:
        seen = {}
        for it in items:
            v = str(it.get(key, "") or "")
            if not v:
                continue
            if v in seen:
                dupes.append({"type": kind, "id": v})
            seen[v] = True

    check_dupes(creatures, "id", "creature")
    check_dupes(weapons, "id", "weapon")
    check_dupes(stages, "id", "stage")
    check_dupes(bosses, "id", "boss")

    creature_ids = {c["id"] for c in creatures if c.get("id")}
    weapon_ids = {w["id"] for w in weapons if w.get("id")}
    boss_ids = {b["id"] for b in bosses if b.get("id")}

    for w in weapons:
        cid = w.get("creatureId")
        if cid and cid not in creature_ids:
            missing_refs.append({"where": "weapons", "id": w.get("id"), "missing": "creatureId", "value": cid})

    for s in stages:
        bid = s.get("bossId")
        if bid and bid not in boss_ids:
            missing_refs.append({"where": "stages", "id": s.get("id"), "missing": "bossId", "value": bid})

    # Evolution prereq node references
    node_ids = {n.get("id") for n in evo_nodes if n.get("id")}
    for pr in evo_prereqs:
        a = pr.get("nodeId") or pr.get("id")
        b = pr.get("requiresNodeId") or pr.get("requires")
        if a and a not in node_ids:
            evo_errs.append({"type": "missing_node", "nodeId": a})
        if b and b not in node_ids:
            evo_errs.append({"type": "missing_prereq", "requiresNodeId": b, "for": a})

    # Asset existence (best-effort): check creature sheet/icon paths if present
    for c in creatures:
        for k in ("alphaSpriteSheetPath", "alphaIconPath"):
            p = c.get(k)
            if not p:
                continue
            disk = os.path.join(assets_root, os.path.normpath(str(p)).replace("assets/", "")) if str(p).startswith("assets/") else os.path.join(assets_root, os.path.normpath(str(p)))
            if not os.path.exists(disk):
                missing_assets.append(str(p))

    return Validation(
        missing_references=missing_refs,
        duplicate_ids=dupes,
        evolution_errors=evo_errs,
        missing_assets=sorted(list(dict.fromkeys(missing_assets))),
    )


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--validate-only", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    launch_rows = read_table(wb["LaunchScope"], id_col_names=("id",))
    launch_by_type: Dict[str, set] = {}
    for r in launch_rows:
        t = str(r.get("scopeType") or "").strip().lower()
        i = str(r.get("id") or "").strip()
        if not t or not i:
            continue
        launch_by_type.setdefault(t, set()).add(i)

    # Creatures
    creature_rows = read_table(wb["Creatures"], id_col_names=("creatureId", "id"))
    creatures: List[Dict[str, Any]] = []
    for r in creature_rows:
        cid = str(r.get("creatureId") or r.get("id") or "").strip()
        if not cid:
            continue
        display = r.get("displayName") or cid
        is_launch = bool(r.get("isLaunch")) or (cid in launch_by_type.get("creature", set()))
        creatures.append({
            "id": cid,
            "displayName": str(display),
            "role": r.get("role"),
            "isLaunch": is_launch,
            "isExperimental": not is_launch,
            "alphaSpriteSheetPath": r.get("alphaSpriteSheetPath"),
            "alphaIconPath": r.get("alphaIconPath"),
            "spriteSource": r.get("spriteSource"),
            # Derivables / compatibility
            "iconId": f"player.{cid}.icon",
            "spriteSheetId": f"player.{cid}.sheet",
        })

    # Weapons
    weapon_rows = read_table(wb["Weapons"], id_col_names=("weaponId", "id"))
    weapons: List[Dict[str, Any]] = []
    for r in weapon_rows:
        wid = str(r.get("weaponId") or r.get("id") or "").strip()
        if not wid:
            continue
        cid = r.get("creatureId")
        is_launch = bool(r.get("isLaunch")) or (wid in launch_by_type.get("weapon", set()))
        weapons.append({
            "id": wid,
            "displayName": r.get("displayName") or wid,
            "creatureId": str(cid).strip() if cid else None,
            "family": r.get("family") or "misc",
            "isLaunch": is_launch,
            "isExperimental": not is_launch,
        })

    # Attach primary weapon id to creatures (best-effort): first matching weapon row
    by_creature_weapon: Dict[str, str] = {}
    for w in weapons:
        if w.get("creatureId") and w.get("id") and w.get("family") == "unique":
            by_creature_weapon.setdefault(w["creatureId"], w["id"])
    for c in creatures:
        c["primaryWeaponId"] = by_creature_weapon.get(c["id"]) or None

    # Weapon families
    families: Dict[str, Dict[str, Any]] = {}
    for w in weapons:
        fam = str(w.get("family") or "misc")
        f = families.setdefault(fam, {"id": fam, "displayName": fam.replace("_", " ").title(), "weaponIds": []})
        f["weaponIds"].append(w["id"])
    weapon_families: List[Dict[str, Any]] = []
    for fam, f in families.items():
        # Experimental if no launch weapons
        has_launch = any((next((w for w in weapons if w["id"] == wid), {}).get("isLaunch") for wid in f["weaponIds"]))
        weapon_families.append({
            "id": f["id"],
            "displayName": f["displayName"],
            "weaponIds": sorted(list(dict.fromkeys(f["weaponIds"]))),
            "experimental": not bool(has_launch),
        })

    # Stages
    stage_rows = read_table(wb["Stages"], id_col_names=("stageId", "id"))
    stages: List[Dict[str, Any]] = []
    for r in stage_rows:
        sid = str(r.get("stageId") or r.get("id") or "").strip()
        if not sid:
            continue
        display = r.get("displayName") or r.get("name") or sid
        is_launch = (sid in launch_by_type.get("stage", set()))
        stages.append({
            "id": sid,
            "displayName": str(display),
            "bossId": r.get("bossId"),
            "bossSpriteSheetId": r.get("bossSpriteSheetId"),
            "isLaunch": is_launch,
            "isExperimental": not is_launch,
            "tone": r.get("tone"),
        })

    # Props
    prop_rows = read_table(wb["Props"], id_col_names=("propId", "id"))
    props: List[Dict[str, Any]] = []
    for r in prop_rows:
        pid = str(r.get("propId") or r.get("id") or "").strip()
        if not pid:
            continue
        is_placeholder = bool(r.get("isPlaceholder")) if "isPlaceholder" in r else False
        props.append({
            "id": pid,
            "displayName": r.get("displayName") or pid,
            "spriteId": r.get("spriteId") or f"prop.{pid}.image",
            "isPlaceholder": is_placeholder,
        })

    # Bosses
    boss_rows = read_table(wb["Bosses"], id_col_names=("bossId", "id"))
    bosses: List[Dict[str, Any]] = []
    for r in boss_rows:
        bid = str(r.get("bossId") or r.get("id") or "").strip()
        if not bid:
            continue
        bosses.append({
            "id": bid,
            "displayName": r.get("displayName") or bid,
            "spriteSheetId": r.get("spriteSheetId") or f"boss.{bid}.sheet",
        })

    # Evolution tables (normalized)
    evo_nodes = read_table(wb["Evo_Nodes"], id_col_names=("id", "nodeId"))
    evo_modules = read_table(wb["Evo_Modules"], id_col_names=("id", "moduleId"))
    evo_visuals = read_table(wb["Evo_Visuals"], id_col_names=("id", "visualId"))
    evo_prereqs = read_table(wb["Evo_Prereqs"], id_col_names=("id", "prereqId"))
    evo_excl = read_table(wb["Evo_Exclusivity"], id_col_names=("id", "groupId"))

    # VFX/SFX
    vfx_rows = read_table(wb["VFX"], id_col_names=("vfxId", "id"))
    sfx_rows = read_table(wb["SFX"], id_col_names=("sfxId", "id"))
    vfx = [{"id": str(r.get("vfxId") or r.get("id")), **{k: r.get(k) for k in r.keys() if k not in ("vfxId",)}} for r in vfx_rows if (r.get("vfxId") or r.get("id"))]
    sfx = [{"id": str(r.get("sfxId") or r.get("id")), **{k: r.get(k) for k in r.keys() if k not in ("sfxId",)}} for r in sfx_rows if (r.get("sfxId") or r.get("id"))]

    # UI tokens
    ui_rows = read_table(wb["UI_Spec"], id_col_names=("Section",))
    palette_row = None
    for r in ui_rows:
        if str(r.get("Section") or "").strip().lower() == "palette lock":
            palette_row = r
            break
    palette_text = str((palette_row or {}).get("Spec") or "")
    tokens = {
        "colors": parse_palette_tokens(palette_text),
        "meta": {"source": os.path.basename(args.xlsx)},
    }
    # HUD layout
    hud_rows = read_table(wb["HUD_Spec"], id_col_names=("Item",))
    hud_coords = None
    for r in hud_rows:
        if str(r.get("Item") or "").strip().lower().startswith("coordinates"):
            hud_coords = str(r.get("Spec") or "")
            break
    hud_layout = parse_hud_boxes(hud_coords or "")

    # Validation
    assets_root = os.path.join(os.path.dirname(os.path.abspath(args.xlsx)), "")
    # In-repo assets folder is relative to cwd; prefer that if present.
    repo_assets = os.path.join(os.path.dirname(__file__), "..", "assets")
    if os.path.isdir(repo_assets):
        assets_root = os.path.abspath(repo_assets)

    v = validate_all(
        creatures=creatures,
        weapons=weapons,
        stages=stages,
        bosses=bosses,
        evo_nodes=[{"id": str(n.get("id") or n.get("nodeId") or "").strip(), **n} for n in evo_nodes],
        evo_prereqs=evo_prereqs,
        assets_root=assets_root,
    )

    report = {
        "schemaVersion": 1,
        "duplicateIds": v.duplicate_ids,
        "missingReferences": v.missing_references,
        "evolutionErrors": v.evolution_errors,
        "missingAssetsOnDisk": v.missing_assets,
    }

    if args.validate_only:
        # Always write validation report so CI/dev has a single artifact.
        write_json(os.path.join(args.out, "validation_report.json"), report)
        return 0

    # Write outputs
    write_json(os.path.join(args.out, "creatures.json"), {"creatures": creatures})
    write_json(os.path.join(args.out, "weapons.json"), {"weapons": weapons})
    write_json(os.path.join(args.out, "weapon_families.json"), {"weaponFamilies": weapon_families})
    write_json(os.path.join(args.out, "stages.json"), {"stages": stages})
    write_json(os.path.join(args.out, "props.json"), {"props": props})
    write_json(os.path.join(args.out, "bosses.json"), {"bosses": bosses})
    write_json(os.path.join(args.out, "vfx.json"), {"vfx": vfx})
    write_json(os.path.join(args.out, "sfx.json"), {"sfx": sfx})
    write_json(os.path.join(args.out, "ui_style_guide.json"), tokens)
    write_json(os.path.join(args.out, "hud_layout.json"), hud_layout)
    write_json(os.path.join(args.out, "validation_report.json"), report)

    evo_dir = os.path.join(args.out, "evolution")
    write_json(os.path.join(evo_dir, "nodes.json"), {"nodes": evo_nodes})
    write_json(os.path.join(evo_dir, "modules.json"), {"modules": evo_modules})
    write_json(os.path.join(evo_dir, "visuals.json"), {"visuals": evo_visuals})
    write_json(os.path.join(evo_dir, "prereqs.json"), {"prereqs": evo_prereqs})
    write_json(os.path.join(evo_dir, "exclusivity.json"), {"exclusivity": evo_excl})

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
